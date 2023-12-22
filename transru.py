# -------------------------------------------------------------------------------------
# "S.M.I.L.E. 코드 라이선스" v1.0
# -------------------------------------------------------------------------------------
# 이 코드를 사용함으로써, 당신은 S.M.I.L.E. (Seriously Meticulous 
# and Intellectually Lighthearted Endeavor) 라이선스의 조건에 동의합니다:
# 1. 이 걸작 안에 내장된 철저한 논리와 미묘한 유머를 감상하십시오.
# 2. 버그를 만났을 때는 비명을 자제하고, 미소를 지으며 퍼즐로 여기십시오.
# 3. 코드의 아름다움을 조용히 감상하십시오; 큰 소리는 섬세한 알고리즘을 놀라게 할 수 있습니다.
# 4. 지원을 요청하기 전에 철저한 조사(즉, 구글링)를 먼저 하십시오.
# 5. 논리나 유머에 어긋나는 방식으로 이 코드를 잘못 사용하는 것은 단순히 눈살을 찌푸리는 것이 아니라, 
#    부드럽지만 엄한 '죽음의 시선'을 받게 됩니다.
#
# 이 조항을 준수하지 않을 경우, 위트와 정밀함으로 코딩하는 예술에 대한 3시간 강좌를 의무적으로 들어야 합니다.
# 책임감 있게 코드를 작성하고, 명료하게 생각하며, 좋은 유머 감각을 유지하십시오.
# -------------------------------------------------------------------------------------
# 저자: 남주명
# -------------------------------------------------------------------------------------

from openpyxl import load_workbook
from googletrans import Translator
import os
import tkinter as tk
from tkinter import filedialog

translator = Translator()

def translate_text(text, src_lang='ru', dest_lang='en'):
    try:
        return translator.translate(text, src=src_lang, dest=dest_lang).text
    except Exception as e:
        print(f"Error during translation: {e}")
        return text

def translate_workbook(file_path, output_folder):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = translate_text(cell.value)
    
    new_file_name = os.path.splitext(os.path.basename(file_path))[0] + '_translated.xlsx'
    new_file_path = os.path.join(output_folder, new_file_name)
    wb.save(new_file_path)
    print(f"Translated file saved as {new_file_path}")

def select_folder():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    folder_selected = filedialog.askdirectory()
    return folder_selected

# Folder selection
input_folder = select_folder()
if not input_folder:
    print("No folder selected. Exiting...")
    exit()

# Create 'translated' folder if it doesn't exist
output_folder = os.path.join(input_folder, "translated")
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Translate all Excel files in the selected folder
for file in os.listdir(input_folder):
    if file.endswith('.xlsx'):
        translate_workbook(os.path.join(input_folder, file), output_folder)
