# -------------------------------------------------------------------------------------
# "S.M.I.L.E. 코드 라이선스" v1.0
# -------------------------------------------------------------------------------------
# 이 코드를 사용함으로써, 당신은 S.M.I.L.E. (Seriously Meticulous 
# and Intellectually Lighthearted Endeavor) 라이선스의 조건에 동의합니다:
# 1. 이 걸작 안에 내장된 철저한 논리와 미묘한 유머를 감상하십시오.
# 2. 버그를 만났을 때는 비명을 자제하고, 미소를 지으며 퍼즐로 여기십시오.
# 3. 코드의 아름다움을 조용히 감상하십시오; 큰 소리는 섬세한 알고리즘을 놀라게 할 수 있습니다.
# 4. 지원을 요청하기 전에 철저한 조사(즉, 구글링)를 먼저 하십시오.
# 5. 논리나 유머에 어긋나는 방식으로 이 코드를 잘못 사용하는 것은 단순히 눈살을 찌푸리는 것이 아니라, 
#    부드럽지만 엄한 '죽음의 시선'을 받게 됩니다.
#
# 이 조항을 준수하지 않을 경우, 위트와 정밀함으로 코딩하는 예술에 대한 3시간 강좌를 의무적으로 들어야 합니다.
# 책임감 있게 코드를 작성하고, 명료하게 생각하며, 좋은 유머 감각을 유지하십시오.
# -------------------------------------------------------------------------------------
# 저자: 남주명
# -------------------------------------------------------------------------------------

import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import filedialog

# Filenames
bs_filename = 'balanced_sheet_raw_ru.xlsx'
is_filename = 'income_statement_raw_ru.xlsx'
cf_filename = 'cash_flow_raw_ru.xlsx'

# Setting up the Tkinter root
root = tk.Tk()
root.withdraw()  # Hide the main window

# Open a dialog to choose a folder
# Select original xls files
folder_path = filedialog.askdirectory()

# Function to read excel data from a given range
def read_excel_data(filename, start_row, start_col, num_rows, num_cols):
    return pd.read_excel(filename, skiprows=start_row-1, usecols=range(start_col-1, start_col-1+num_cols), nrows=num_rows)

# Pixels to cell width
def pixels_to_width(pixels):
    return pixels / (7 + 1/4)
desired_pixel_width = 100
excel_width = pixels_to_width(desired_pixel_width)

# Searching for filenames
balanced_sheet_1 = None
balanced_sheet_2 = None
income_statement_1 = None
income_statement_2 = None
cash_flow_data_1 = None
cash_flow_data_2 = None

# Check if a folder was selected
if folder_path:
    for file in os.listdir(folder_path):
        if "баланс. Лист 1" in file and file.endswith('.xls'):
            balanced_sheet_1 = os.path.join(folder_path, file)
        elif "баланс. Лист 2" in file and file.endswith('.xls'):
            balanced_sheet_2 = os.path.join(folder_path, file)
        elif "результатах. Лист 1" in file and file.endswith('.xls'):
            income_statement_1 = os.path.join(folder_path, file)
        elif "результатах. Лист 2" in file and file.endswith('.xls'):
            income_statement_2 = os.path.join(folder_path, file)
        elif "средств. Лист 1" in file and file.endswith('.xls'):
            cash_flow_data_1 = os.path.join(folder_path, file)
        elif "средств. Лист 2" in file and file.endswith('.xls'):
            cash_flow_data_2 = os.path.join(folder_path, file)

# Reading the data from the first bs file
if balanced_sheet_1:
    bs_data_1 = read_excel_data(balanced_sheet_1, 28, 3, 59, 33)
    bs_data_1 = bs_data_1.dropna(how='all', axis=0).dropna(how='all', axis=1)
else:
    bs_data_1 = "No file with 'баланс. Лист 1' in the name found in the current directory."

# Reading the data from the second bs file
if balanced_sheet_2:
    bs_data_2 = read_excel_data(balanced_sheet_2, 2, 3, 52, 15)
    bs_data_2 = bs_data_2.dropna(how='all', axis=0).dropna(how='all', axis=1)
else:
    bs_data_2 = "No file with 'баланс. Лист 2' in the name found in the current directory."

# Reading the data from the first is file
if income_statement_1:
    is_data_1 = read_excel_data(income_statement_1, 15, 3, 18, 16)
    is_data_1 = is_data_1.dropna(how='all', axis=0).dropna(how='all', axis=1)
else:
    is_data_1 = "No file with 'результатах. Лист 1' in the name found in the current directory."
# Leave the second column's title blank
is_data_1.columns.values[1] = ''

# Reading the data from the second is file
if income_statement_2:
    is_data_2 = read_excel_data(income_statement_2, 2, 3, 50, 15)
    is_data_2 = is_data_2.dropna(how='all', axis=0).dropna(how='all', axis=1)
else:
    is_data_2 = "No file with 'результатах. Лист 2' in the name found in the current directory."
# Add an empty column at the second position with an empty string as the column name
is_data_2.insert(1, '', pd.NA)

# Reading the data from the first cf file
if cash_flow_data_1:
    cf_data_1 = read_excel_data(cash_flow_data_1, 13, 3, 31, 13)
    cf_data_1 = cf_data_1.dropna(how='all', axis=0).dropna(how='all', axis=1)
else:
    cf_data_1 = "No file with 'средств. Лист 1' in the name found in the current directory."

# Reading the data from the second cf file
if cash_flow_data_2:
    cf_data_2 = read_excel_data(cash_flow_data_2, 2, 3, 45, 13)
    cf_data_2 = cf_data_2.dropna(how='all', axis=0).dropna(how='all', axis=1)
else:
    cf_data_2 = "No file with 'средств. Лист 2' in the name found in the current directory."


# Concatenating the two dataframes
bs_merged_data = pd.concat([bs_data_1, bs_data_2], ignore_index=True)
# Leave the second column's title blank
bs_merged_data.columns.values[1] = ''
# Dropping the merged data to an Excel file
bs_merged_data.to_excel(bs_filename, index=False)

# Concatenating the two dataframes
is_merged_data = pd.concat([is_data_1, is_data_2], ignore_index=True)
# Dropping the merged data to an Excel file
is_merged_data.to_excel(is_filename, index=False)

# Concatenating the two dataframes
cf_merged_data = pd.concat([cf_data_1, cf_data_2], ignore_index=True)
# Leave the second column's title blank
cf_merged_data.columns.values[1] = ''
# Dropping the merged data to an Excel file
cf_merged_data.to_excel(cf_filename, index=False)


# num_format = '#,##0,;-#,##0,,"M";-'
# num_format = '_-* #,##0_-;-* #,##0_-;_-* "-"_-;_-@_-'
# num_format = '#,##0_);[빨강](#,##0)'
# num_format = '#,##0_);(#,##0)'
num_format = '#,##0,;(#,##0,)'


# Open the BS Excel file for editing
workbook = openpyxl.load_workbook(bs_filename)
sheet = workbook.active

# Define the fill color for rows where the first column is 'BALANCE'
balance_fill = PatternFill(start_color='8FBAC8', end_color='8FBAC8', fill_type='solid')
total_section_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

# Iterate over the rows for conditional formatting
for row in sheet.iter_rows(min_row=2):  # Start from the second row to skip the header
    # Hide rows where the last three columns are all zeros
    if all(cell.value == 0 or cell.value == '-' or cell.value == None for cell in row[-3:]):  # Check if the last three cells in the row are all trivial
        sheet.row_dimensions[row[0].row].hidden = True

    # Change row color if the first column is 'BALANCE'
    if row[0].value == 'БАЛАНС':
        for cell in row:
            cell.fill = balance_fill
        # sheet.row_dimensions[row[0].row].hidden = True

    # Change row color if the first column contains 'Total section'
    if 'Итого по разделу' in str(row[0].value):
        for cell in row:
            cell.fill = total_section_fill
        # 당장 뭐에 쓸지 모르겠으니 일단 숨기고 봄
        sheet.row_dimensions[row[0].row].hidden = True

# Apply the format to all cells in the last three columns
for row in sheet.iter_rows(min_col=sheet.max_column - 2, max_col=sheet.max_column):
    for cell in row:
        cell.number_format = num_format

# Apply the width to the last three columns
for i in range(3):
    col_letter = openpyxl.utils.cell.get_column_letter(sheet.max_column - i)
    sheet.column_dimensions[col_letter].width = excel_width

# Iterate over the rows and hide rows where the last two columns are all zeros
for row in sheet.iter_rows(min_row=2):  # Start from the second row to skip the header
    if all(cell.value == 0 or cell.value == '-' or cell.value == None for cell in row[-2:]):  # Check if the last two cells in the row are all trivial
        sheet.row_dimensions[row[0].row].hidden = True  # Hide the row

# Save the changes to the Excel file
workbook.save(bs_filename)


# Open the IS file for editing
workbook = openpyxl.load_workbook(is_filename)
sheet = workbook.active

# Apply the format to all cells in the last two columns
for row in sheet.iter_rows(min_col=sheet.max_column - 1, max_col=sheet.max_column):
    for cell in row:
        cell.number_format = num_format

# Apply the width to the last two columns
for i in range(2):
    col_letter = openpyxl.utils.cell.get_column_letter(sheet.max_column - i)
    sheet.column_dimensions[col_letter].width = excel_width

# Iterate over the rows and hide rows where the last two columns are all zeros
for row in sheet.iter_rows(min_row=2):  # Start from the second row to skip the header
    if all(cell.value == 0 or cell.value == '-' or cell.value == None for cell in row[-2:]):  # Check if the last two cells in the row are all trivial
        sheet.row_dimensions[row[0].row].hidden = True  # Hide the row

# Save the modified Excel file
workbook.save(is_filename)


# Open the CF file for editing
workbook = openpyxl.load_workbook(cf_filename)
sheet = workbook.active

# Iterate through the DataFrame and hide rows in the Excel sheet
# Skip the header row by starting from index 2 (row 1 in openpyxl)
for i in cf_merged_data.index:
    # Check if the last column is empty or zero
    if pd.isna(cf_merged_data.iloc[i, -1]) or cf_merged_data.iloc[i, -1] == 0:
        sheet.row_dimensions[i + 2].hidden = True

# Hide rows with indentation
for row in sheet.iter_rows(min_row=2): # Start from the second row to skip the header
    # Hide rows where if first item is empty
    if row[0].value == None:  # Check if the first item is empty
        sheet.row_dimensions[row[0].row].hidden = True

# Apply the format to all cells in the last two columns
for row in sheet.iter_rows(min_col=sheet.max_column - 1, max_col=sheet.max_column):
    for cell in row:
        cell.number_format = num_format

# Apply the width to the last two columns
for i in range(2):
    col_letter = openpyxl.utils.cell.get_column_letter(sheet.max_column - i)
    sheet.column_dimensions[col_letter].width = excel_width

# Save the modified Excel file
workbook.save(cf_filename)

# print(bs_merged_data)
# print(is_merged_data)
# print(cf_merged_data)
print("Merged data saved")


# Translate the output files
from openpyxl import load_workbook
from googletrans import Translator

translator = Translator()

def translate_text(text, src_lang='ru', dest_lang='en'):
    try:
        return translator.translate(text, src=src_lang, dest=dest_lang).text
    except Exception as e:
        print(f"Error during translation: {e}")
        return text

def translate_workbook(file_path, output_folder):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = translate_text(cell.value)
    
    new_file_name = os.path.splitext(os.path.basename(file_path))[0] + '_translated.xlsx'
    new_file_path = os.path.join(output_folder, new_file_name)
    wb.save(new_file_path)
    print(f"Translated file saved as {new_file_path}")

# Translate all Excel files in the selected folder
for file in os.listdir(os.getcwd()):
    if file.endswith('.xlsx'):
        print("Translating " + file)
        translate_workbook(os.path.join(os.getcwd(), file), os.getcwd())
